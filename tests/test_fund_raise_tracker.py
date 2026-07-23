"""fund_raise_tracker 的单元测试：验证 xlsx 能生成且内容正确。"""
import openpyxl

import fund_raise_tracker as frt


def test_build_creates_xlsx(tmp_path, monkeypatch):
    out = tmp_path / "reports" / "tracker.xlsx"
    monkeypatch.setattr(frt, "OUT", str(out))
    frt.build()
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    ws = wb.active
    assert ws.title == "募集进度"
    # 表头 + 3 行示例数据
    assert ws.max_row == 4
    headers = [c.value for c in ws[1]]
    assert headers == ["LP 名称", "承诺金额(万)", "已实缴(万)", "进度%", "状态"]


def test_progress_percentage_is_correct(tmp_path, monkeypatch):
    out = tmp_path / "reports" / "tracker.xlsx"
    monkeypatch.setattr(frt, "OUT", str(out))
    frt.build()
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # 家族办公室B：承诺 3000 已缴 3000 → 100%
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, commit, paid, pct, status = row
        assert pct == round(paid / commit * 100, 1)
