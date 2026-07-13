"""
基金募集进度跟踪器 v2026.1
生成 xlsx：LP 名单、承诺金额、已实缴、进度%。
依赖：openpyxl
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Windows 控制台默认非 UTF-8，中文输出会 UnicodeEncodeError
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

OUT = "reports/fund_raise_tracker.xlsx"


def build():
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "募集进度"
    headers = ["LP 名称", "承诺金额(万)", "已实缴(万)", "进度%", "状态"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    sample = [
        ("产业基金A", 5000, 2000, "进行中"),
        ("家族办公室B", 3000, 3000, "已完成"),
        ("高净值客户C", 2000, 500, "进行中"),
    ]
    for name, commit, paid, status in sample:
        pct = round(paid / commit * 100, 1)
        ws.append([name, commit, paid, pct, status])
    wb.save(OUT)
    print(f"已生成 {OUT}")


if __name__ == "__main__":
    build()
